import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
import sys
import datetime
import tkinter as tk
from tkcalendar import Calendar
from tkinter.filedialog import askopenfilename, asksaveasfilename
import babel.numbers

def resource_path(relative_path):
    absolute_path = os.path.abspath(__file__)
    root_path = os.path.dirname(absolute_path)
    base_path = getattr(sys, '_MEIPASS', root_path)
    return os.path.join(base_path, relative_path)

#first row has 5 dates, tues-sat
#second row is under first row

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Week 1"

inc_delta = datetime.timedelta(days=1)
week_delta = datetime.timedelta(days=7)

colorList = ["4577C7", "CF7977", "A2D88C", "B1A0C7", "FABF8F"]
            #  blue     red       green    purple      orange

def writeDayHeaders(start_date, sheet):
    headers = []
    # store dates from start time
    for i in range(5):
        headers.append(start_date.strftime("%A, %B %d, %Y"))
        start_date += inc_delta
    # Write headers
    column = 1
    header_num = 0
    # Fill in color for each column
    for i in range(5):
        sheet.cell(row=1, column=column, value=headers[header_num]).fill = PatternFill(start_color=colorList[i]
                                                                                    , end_color=colorList[i], fill_type='solid')
        sheet.cell(row=1, column=column).font = Font(bold=True)
        column += 4
        header_num += 1
    start_cell = "A1"
    end_cell = "D1"


    # Merge Day columns
    for i in range(5):
            sheet.merge_cells(f"{start_cell}:{end_cell}")
            start_cell = chr(ord(start_cell[0])+4) + "1"
            end_cell = chr(ord(end_cell[0])+4) + "1"
            
    # Set width of name columns
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['F'].width = 22
    sheet.column_dimensions['J'].width = 22
    sheet.column_dimensions['N'].width = 22
    sheet.column_dimensions['R'].width = 22

def writeSubheaders(sheet):
    sub_column_data = ["Arrival", "Name", "# in Party", "Reason",
                   "Arrival", "Name", "# in Party", "Reason",
                   "Arrival", "Name", "# in Party", "Reason",
                   "Arrival", "Name", "# in Party", "Reason",
                   "Arrival", "Name", "# in Party", "Reason"]

    sheet.append(sub_column_data)
    
def colorSubheaders(sheet):
    for cell in sheet[2]:
        cell.fill = PatternFill(start_color='FEE198'
                                , end_color='FEE198', fill_type='solid')
        cell.font = Font(bold=True)

def saveWorkbook():
    file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        workbook.save(file_path)
        
def update_label_selected_date(event):
    selected_date = cal.get_date()
    selected_date_label.config(text=f"Selected Date: {selected_date}", fg="black")

def is_tuesday(date):
    d_date = datetime.datetime.strptime(date, "%m/%d/%y")
    return d_date.weekday() == 1

def generate_sheet():
    selected_date = cal.get_date()
    if is_tuesday(selected_date) != 1:
        selected_date_label.config(text=f"Selected Date: {selected_date} is not a Tuesday", fg="red")
        return
    
    
    start_date = datetime.datetime.strptime(selected_date, "%m/%d/%y")
    
    writeDayHeaders(start_date, sheet)
    writeSubheaders(sheet)
    colorSubheaders(sheet)
    start_date += week_delta
    for week_num in range(2, 11):
        sheet_name = f"Week {week_num}"
        new_sheet = workbook.create_sheet(title=sheet_name)
        writeDayHeaders(start_date, new_sheet)
        writeSubheaders(new_sheet)
        colorSubheaders(new_sheet)
        start_date += week_delta
    saveWorkbook()
    

if __name__ == "__main__":
    window = tk.Tk()
    window.title("FRC Spreadsheet Generator")


    # Top blue bar
    frame1 = tk.Frame(master=window, height=20, bg="#08538c")
    uciLogo = tk.PhotoImage(file=resource_path("UCILogo.png"))
    lbl_logo = tk.Label(master=frame1,width=250, height=60, image=uciLogo, background="#08538c")
    lbl_logo.grid(row=0, column=0)

    tk.Label(master=frame1, text="FRC SPREADSHEET GENERATOR", fg="white", bg="#08538c", font=("Arial", 16, "bold")).grid(row=0, column=1, padx= 10)
    frame1.pack(fill=tk.X)


    # Calendar
    frame2 = tk.Frame(master=window, pady=5)
    
    # left side of grid
    cal = Calendar(master=frame2, selectmode="day", showweeknumbers=False)
    cal.grid(row=0, column=0, padx=20, pady=10)
    cal.bind("<<CalendarSelected>>", update_label_selected_date)

    selected_date_label = tk.Label(master=frame2, text="Selected Date: ")
    selected_date_label.grid(row=1, column=0, pady=5)
    
    # right side of grid
    generate_frame = tk.Frame(master=frame2)
    generate_frame.grid(row=0, column=1, rowspan=2, sticky="nsew")
    
    
    readme = tk.Label(master=generate_frame, text="            Instructions            ", font=("Arial", 16, "bold"), bg="#AAACAD")
    readme.pack(pady=5, padx=20)
    
    bullet_points = "1. Select a Tuesday on the Calendar\n2. Click Generate Spreadsheet\n3. Choose where to save file\n4. Open the excel file using google sheets"
    lbl_instr = tk.Label(master=generate_frame, 
                         text=bullet_points, justify="left")
    lbl_instr.pack()
    
    btn_gen = tk.Button(master=generate_frame, text="Generate Spreadsheet", command=generate_sheet)
    btn_gen.pack(side="bottom", pady=20, padx=20)

    
    # Configure grid_rowconfigure and grid_columnconfigure to make generate_frame expand
    frame2.grid_rowconfigure(0, weight=1)
    frame2.grid_rowconfigure(1, weight=1)
    frame2.grid_columnconfigure(1, weight=1)
        
    frame2.pack(fill=tk.X)

    window.mainloop()